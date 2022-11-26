from openpyxl import load_workbook
import openpyxl

if __name__ == '__main__':
    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active

    ws['A1'] = '(x,y)'
    ws['B1'] = '强度'

    # 1.打开 Excel 表格并获取表格名称
    workbook1 = load_workbook(filename="X-Wave-0.5.xlsx")
    workbook2 = load_workbook(filename="Y-Sn-0.5.xlsx")
    # 2.通过 sheet 名称获取表格
    sheet1 = workbook1["Sheet1"]
    sheet2 = workbook2["Sheet1"]
    # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
    print('sheet1.dimensions:'+sheet1.dimensions)
    print('sheet2.dimensions:'+sheet2.dimensions)
    # 4.2sheet.cell(row=, column=)方式
    # 5. 获取一系列格子
    # 获取 A1:C2 区域的值

    cell1 = sheet1["A2:B169"]
    cell2 = sheet2["A2:B1025"]
    c=1
    for i in cell1:
        r = 2
        for j in cell2:
            # ws.cell(row=r, column=c, value=str(i[0].value)+','+str(j[0].value))
            ws.cell(row=r, column=c, value=i[1].value*j[1].value)
            r=r+1
        c=c+1
    wb.save('result-0.5.xlsx')
    workbook1.save('X-Wave-0.5.xlsx')
    workbook2.save('Y-Sn-0.5.xlsx')
    print("end")
